"""
Generate Excel spreadsheet showing SharePoint read/write operations for each action.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Define the data structure
data = {
    'Action': [
        'Run Property Audit',
        'Portfolio from Home',
        'Back to Home',
        'Run Single Lease Audit',
        'Run Bulk Audit',
        'Portfolio from Lease Screen',
        '',  # Separator
        'Property View',
        'Lease View',
    ],
    'AuditRuns2': [
        'Write-yes\nRead-no',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        'Write-yes\nRead-no',
        'Write-yes\nRead-no',
        'Write-n/a\nRead-no',
        '',
        'Write-n/a\nRead-yes',
        'Write-n/a\nRead-yes',
    ],
    'RunDisplaySnapshots': [
        'Write-yes\nRead-no',
        'Write-n/a\nRead-yes',
        'Write-n/a\nRead-no',
        'Write-yes\nRead-no',
        'Write-yes\nRead-no',
        'Write-n/a\nRead-yes',
        '',
        'Write-n/a\nRead-yes',
        'Write-n/a\nRead-yes',
    ],
    'ExceptionMonths': [
        'n/a\nRead-no',
        'Write-n/a\nRead-yes',
        'Write-n/a\nRead-no',
        'n/a\nRead-no',
        'n/a\nRead-no',
        'Write-n/a\nRead-yes',
        '',
        'Write-n/a\nRead-yes',
        'Write-n/a\nRead-yes',
    ],
    'LeaseTerms': [
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        '',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
    ],
    'LeaseTermsEvidence': [
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        '',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
    ],
    'LeaseTermsSet': [
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
        '',
        'Write-n/a\nRead-no',
        'Write-n/a\nRead-no',
    ],
    'AuditRunMetrics': [
        'Write-yes\nRead-minimal*',
        'Write-n/a\nRead-yes',
        'Write-n/a\nRead-yes',
        'Write-yes\nRead-minimal*',
        'Write-yes\nRead-minimal*',
        'Write-n/a\nRead-yes',
        '',
        'Write-n/a\nRead-yes',
        'Write-n/a\nRead-yes',
    ],
}

# Create DataFrame
df = pd.DataFrame(data)

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "SharePoint Operations"

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
separator_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
action_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
action_font = Font(bold=True, size=10)
view_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
view_font = Font(bold=True, italic=True, size=10)
write_yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
read_yes_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
na_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Write headers
for col_idx, col_name in enumerate(df.columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Write data
for row_idx, row in enumerate(df.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        
        # Apply special formatting
        action_value = row[0]  # First column is Action
        
        # Separator row
        if action_value == '':
            cell.fill = separator_fill
        # View rows (Property View, Lease View)
        elif 'View' in action_value:
            if col_idx == 1:  # Action column
                cell.fill = view_fill
                cell.font = view_font
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        # Regular action rows
        else:
            if col_idx == 1:  # Action column
                cell.fill = action_fill
                cell.font = action_font
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                # Color code based on content
                if value:
                    if 'Write-yes' in value:
                        cell.fill = write_yes_fill
                    elif 'Read-yes' in value:
                        cell.fill = read_yes_fill
                    elif 'n/a' in value:
                        cell.fill = na_fill
                    elif 'Read-no' in value or 'Write-n/a' in value:
                        # Light gray for no operations
                        pass

# Set column widths
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 20
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 20

# Set row heights for better visibility
for row in range(2, len(df) + 2):
    ws.row_dimensions[row].height = 35

ws.row_dimensions[1].height = 30  # Header row

# Add notes section
notes_start_row = len(df) + 4
ws.cell(row=notes_start_row, column=1, value="NOTES:").font = Font(bold=True, size=11)
ws.cell(row=notes_start_row + 1, column=1, value="• Write-yes = List is written to during this operation")
ws.cell(row=notes_start_row + 2, column=1, value="• Read-yes = List is read from during this operation")
ws.cell(row=notes_start_row + 3, column=1, value="• Write-n/a = List is not written to (operation is read-only or doesn't trigger writes)")
ws.cell(row=notes_start_row + 4, column=1, value="• Read-no = List is not read from")
ws.cell(row=notes_start_row + 5, column=1, value="• n/a = Not applicable (e.g., ExceptionMonths only written on manual resolution)")
ws.cell(row=notes_start_row + 6, column=1, value="• Read-minimal* = Only reads list metadata (run IDs, list structure), not actual audit data")

for row in range(notes_start_row + 1, notes_start_row + 7):
    ws.cell(row=row, column=1).font = Font(size=9)

# Add operation details section
details_start_row = notes_start_row + 9
ws.cell(row=details_start_row, column=1, value="KEY FINDINGS:").font = Font(bold=True, size=11)
ws.cell(row=details_start_row + 1, column=1, value="1. All audit operations (Property, Single Lease, Bulk) write to: AuditRuns2, RunDisplaySnapshots, and Audit Run Metrics")
ws.cell(row=details_start_row + 2, column=1, value="2. All portfolio views read from: RunDisplaySnapshots, Audit Run Metrics, and ExceptionMonths")
ws.cell(row=details_start_row + 3, column=1, value="3. ExceptionMonths is only written during manual exception resolution (separate user action)")
ws.cell(row=details_start_row + 4, column=1, value="4. Lease Term lists are separate workflows not triggered by regular audit operations")
ws.cell(row=details_start_row + 5, column=1, value="5. Navigation-only actions (Portfolio from Home, Back to Home) are read-only display operations")

for row in range(details_start_row + 1, details_start_row + 6):
    cell = ws.cell(row=row, column=1)
    cell.font = Font(size=9)
    cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

# Save workbook
output_file = 'SharePoint_Operations_Matrix.xlsx'
wb.save(output_file)
print(f"✅ Excel file created: {output_file}")
print(f"   Location: z:\\Shared\\Technology\\AI Projects\\LeaseFileAudit\\{output_file}")
